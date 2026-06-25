from __future__ import annotations

import json
from io import BytesIO
from typing import Any

import httpx
from openpyxl import load_workbook

from fastapi.testclient import TestClient
from starlette.datastructures import Headers

from app import main
from app.main import app

client = TestClient(app)


def test_health_returns_unified_json() -> None:
    response = client.get("/api/health")

    assert response.status_code == 200
    assert response.json() == {
        "code": 0,
        "message": "ok",
        "data": {"status": "healthy", "service": "ai-extractor-backend"},
    }


def test_parse_text_and_file_returns_structured_payload() -> None:
    response = client.post(
        "/api/parse",
        data={"text": "合同标题\n金额：100"},
        files={"files": ("contract.txt", b"hello", "text/plain")},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["code"] == 0
    assert body["message"] == "ok"
    assert body["data"]["result"]["summary"] == "合同标题\n金额：100"
    assert body["data"]["result"]["source"]["file_count"] == 1
    assert body["data"]["result"]["source"]["files"][0]["filename"] == "contract.txt"


def test_parse_image_without_vision_api_returns_standard_json(monkeypatch: Any) -> None:
    monkeypatch.delenv(main.LLM_BASE_URL_ENV, raising=False)

    response = client.post(
        "/api/parse",
        data={"text": "请抽取图片"},
        files={"files": ("form.png", b"fake-image", "image/png")},
    )

    assert response.status_code == 200
    result = response.json()["data"]["result"]
    assert set(result) == {"document_info", "sections", "raw_text", "warnings"}
    assert result["document_info"]["confidence"] == 0
    assert result["sections"][0]["fields"][0] == {
        "field_name": "form.png",
        "field_value": "",
        "status": "uncertain",
        "source_hint": "uploaded image: form.png",
    }
    assert main.LLM_BASE_URL_ENV in result["warnings"][0]


def test_parse_image_uses_provider_neutral_json_result(monkeypatch: Any) -> None:
    async def fake_call_vision_model_api(text: str, image_files: list[dict[str, Any]]) -> dict[str, Any]:
        assert text == "识别表格"
        assert image_files[0]["filename"] == "table.jpg"
        assert image_files[0]["content"] == b"image-bytes"
        return main.normalize_extraction_result(
            {
                "document_info": {"title": "巡检表", "id": "A-1", "confidence": 0.87},
                "sections": [
                    {
                        "section_name": "基础信息",
                        "fields": [
                            {
                                "field_name": "温度",
                                "field_value": "23 °C",
                                "status": "filled",
                                "source_hint": "表格第一行",
                            },
                            {
                                "field_name": "备注",
                                "field_value": "",
                                "status": "empty",
                                "source_hint": "底部备注栏",
                            },
                        ],
                    }
                ],
                "raw_text": "温度 23 °C",
                "warnings": [],
            }
        )

    monkeypatch.setattr(main, "call_vision_model_api", fake_call_vision_model_api)

    response = client.post(
        "/api/parse",
        data={"text": "识别表格"},
        files={"files": ("table.jpg", b"image-bytes", "image/jpeg")},
    )

    assert response.status_code == 200
    result = response.json()["data"]["result"]
    assert result["document_info"] == {"title": "巡检表", "id": "A-1", "confidence": 0.87}
    assert result["sections"][0]["fields"][0]["status"] == "filled"
    assert result["sections"][0]["fields"][1]["status"] == "empty"
    assert result["raw_text"] == "温度 23 °C"


def test_vision_api_reads_llm_environment_variables(monkeypatch: Any) -> None:
    captured: dict[str, Any] = {}

    class FakeResponse:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict[str, Any]:
            return {
                "document_info": {"title": "测试表", "id": "T-1", "confidence": 0.9},
                "sections": [
                    {
                        "section_name": "结果",
                        "fields": [
                            {
                                "field_name": "状态",
                                "field_value": "通过",
                                "status": "filled",
                                "source_hint": "图片右上角",
                            }
                        ],
                    }
                ],
                "raw_text": "状态 通过",
                "warnings": [],
            }

    class FakeAsyncClient:
        def __init__(self, timeout: float) -> None:
            captured["timeout"] = timeout

        async def __aenter__(self) -> "FakeAsyncClient":
            return self

        async def __aexit__(self, exc_type: Any, exc: Any, traceback: Any) -> None:
            return None

        async def post(self, url: str, headers: dict[str, str], json: dict[str, Any]) -> FakeResponse:
            captured["url"] = url
            captured["headers"] = headers
            captured["json"] = json
            return FakeResponse()

    monkeypatch.setenv(main.LLM_BASE_URL_ENV, "https://llm.example.com/extract")
    monkeypatch.setenv(main.LLM_API_KEY_ENV, "test-key")
    monkeypatch.setenv(main.LLM_MODEL_ENV, "vision-model-test")
    monkeypatch.setattr(main.httpx, "AsyncClient", FakeAsyncClient)

    response = client.post(
        "/api/parse",
        data={"text": "读取图片"},
        files={"files": ("form.png", b"image-bytes", "image/png")},
    )

    assert response.status_code == 200
    assert captured["timeout"] == main.DEFAULT_VISION_API_TIMEOUT_SECONDS
    assert captured["url"] == "https://llm.example.com/extract"
    assert captured["headers"]["Authorization"] == "Bearer test-key"
    assert captured["json"]["model"] == "vision-model-test"
    assert captured["json"]["text"] == "读取图片"
    assert captured["json"]["images"][0]["base64"] == "aW1hZ2UtYnl0ZXM="
    assert response.json()["data"]["result"]["document_info"]["title"] == "测试表"


def test_messages_api_url_uses_messages_payload_and_parses_content(monkeypatch: Any) -> None:
    captured: dict[str, Any] = {}

    class FakeResponse:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict[str, Any]:
            return {
                "content": [
                    {
                        "type": "text",
                        "text": json.dumps(
                            {
                                "document_info": {"title": "消息接口表", "id": "M-1", "confidence": 0.8},
                                "sections": [],
                                "raw_text": "消息接口表",
                                "warnings": [],
                            },
                            ensure_ascii=False,
                        ),
                    }
                ]
            }

    class FakeAsyncClient:
        def __init__(self, timeout: float) -> None:
            return None

        async def __aenter__(self) -> "FakeAsyncClient":
            return self

        async def __aexit__(self, exc_type: Any, exc: Any, traceback: Any) -> None:
            return None

        async def post(self, url: str, headers: dict[str, str], json: dict[str, Any]) -> FakeResponse:
            captured["url"] = url
            captured["headers"] = headers
            captured["json"] = json
            return FakeResponse()

    monkeypatch.setenv(main.LLM_BASE_URL_ENV, "https://api.vveai.com/v1/messages")
    monkeypatch.setenv(main.LLM_API_KEY_ENV, "test-key")
    monkeypatch.setenv(main.LLM_MODEL_ENV, "claude-compatible-vision")
    monkeypatch.setattr(main.httpx, "AsyncClient", FakeAsyncClient)

    response = client.post(
        "/api/parse",
        data={"text": "用户备注"},
        files={"files": ("form.png", b"image-bytes", "image/png")},
    )

    assert response.status_code == 200
    assert captured["url"] == "https://api.vveai.com/v1/messages"
    assert captured["headers"]["anthropic-version"] == "2023-06-01"
    assert captured["json"]["model"] == "claude-compatible-vision"
    assert captured["json"]["max_tokens"] == 4096
    assert captured["json"]["messages"][0]["role"] == "user"
    content = captured["json"]["messages"][0]["content"]
    assert content[0]["type"] == "text"
    assert "用户备注" in content[0]["text"]
    assert content[1]["type"] == "image"
    assert content[1]["source"] == {
        "type": "base64",
        "media_type": "image/png",
        "data": "aW1hZ2UtYnl0ZXM=",
    }
    assert response.json()["data"]["result"]["document_info"]["title"] == "消息接口表"


def test_chat_completions_url_uses_openai_compatible_messages_payload(monkeypatch: Any) -> None:
    captured: dict[str, Any] = {}

    class FakeResponse:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict[str, Any]:
            return {
                "choices": [
                    {
                        "message": {
                            "content": json.dumps(
                                {
                                    "document_info": {"title": "Chat Completions 表", "id": "C-1", "confidence": 0.91},
                                    "sections": [],
                                    "raw_text": "Chat Completions 表",
                                    "warnings": [],
                                },
                                ensure_ascii=False,
                            )
                        }
                    }
                ]
            }

    class FakeAsyncClient:
        def __init__(self, timeout: float) -> None:
            return None

        async def __aenter__(self) -> "FakeAsyncClient":
            return self

        async def __aexit__(self, exc_type: Any, exc: Any, traceback: Any) -> None:
            return None

        async def post(self, url: str, headers: dict[str, str], json: dict[str, Any]) -> FakeResponse:
            captured["url"] = url
            captured["headers"] = headers
            captured["json"] = json
            return FakeResponse()

    monkeypatch.setenv(main.LLM_BASE_URL_ENV, "https://api.quickrouter.ai/v1/chat/completions")
    monkeypatch.setenv(main.LLM_API_KEY_ENV, "test-key")
    monkeypatch.setenv(main.LLM_MODEL_ENV, "gpt-5.4-nano")
    monkeypatch.setattr(main.httpx, "AsyncClient", FakeAsyncClient)

    response = client.post(
        "/api/parse",
        data={"text": "用户备注"},
        files={"files": ("form.png", b"image-bytes", "image/png")},
    )

    assert response.status_code == 200
    assert captured["url"] == "https://api.quickrouter.ai/v1/chat/completions"
    assert "anthropic-version" not in captured["headers"]
    assert captured["headers"]["Authorization"] == "Bearer test-key"
    assert captured["json"]["model"] == "gpt-5.4-nano"
    assert captured["json"]["response_format"] == {"type": "json_object"}
    content = captured["json"]["messages"][0]["content"]
    assert content[0]["type"] == "text"
    assert "用户备注" in content[0]["text"]
    assert content[1] == {
        "type": "image_url",
        "image_url": {"url": "data:image/png;base64,aW1hZ2UtYnl0ZXM="},
    }
    assert response.json()["data"]["result"]["document_info"]["title"] == "Chat Completions 表"


def test_vision_api_http_error_includes_response_body(monkeypatch: Any) -> None:
    class FakeResponse:
        status_code = 400
        text = "missing required field: messages"

        def raise_for_status(self) -> None:
            request = httpx.Request("POST", "https://api.vveai.com/v1/messages")
            response = httpx.Response(self.status_code, request=request, text=self.text)
            raise httpx.HTTPStatusError("Bad Request", request=request, response=response)

    class FakeAsyncClient:
        def __init__(self, timeout: float) -> None:
            return None

        async def __aenter__(self) -> "FakeAsyncClient":
            return self

        async def __aexit__(self, exc_type: Any, exc: Any, traceback: Any) -> None:
            return None

        async def post(self, url: str, headers: dict[str, str], json: dict[str, Any]) -> FakeResponse:
            return FakeResponse()

    monkeypatch.setenv(main.LLM_BASE_URL_ENV, "https://api.vveai.com/v1/messages")
    monkeypatch.setenv(main.LLM_MODEL_ENV, "claude-compatible-vision")
    monkeypatch.setattr(main.httpx, "AsyncClient", FakeAsyncClient)

    response = client.post(
        "/api/parse",
        data={"text": "读取图片"},
        files={"files": ("form.png", b"image-bytes", "image/png")},
    )

    assert response.status_code == 200
    warning = response.json()["data"]["result"]["warnings"][0]
    assert "status=400" in warning
    assert "missing required field: messages" in warning


def test_vision_api_non_json_response_returns_warning(monkeypatch: Any) -> None:
    class FakeResponse:
        status_code = 200
        text = "not json"

        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict[str, Any]:
            raise json.JSONDecodeError("Expecting value", self.text, 0)

    class FakeAsyncClient:
        def __init__(self, timeout: float) -> None:
            return None

        async def __aenter__(self) -> "FakeAsyncClient":
            return self

        async def __aexit__(self, exc_type: Any, exc: Any, traceback: Any) -> None:
            return None

        async def post(self, url: str, headers: dict[str, str], json: dict[str, Any]) -> FakeResponse:
            return FakeResponse()

    monkeypatch.setenv(main.LLM_BASE_URL_ENV, "https://llm.example.com/extract")
    monkeypatch.setenv(main.LLM_MODEL_ENV, "vision-model-test")
    monkeypatch.setattr(main.httpx, "AsyncClient", FakeAsyncClient)

    response = client.post(
        "/api/parse",
        data={"text": "读取图片"},
        files={"files": ("form.png", b"image-bytes", "image/png")},
    )

    assert response.status_code == 200
    result = response.json()["data"]["result"]
    assert result["sections"][0]["fields"][0]["status"] == "uncertain"
    assert "not valid JSON" in result["warnings"][0]
    assert "not json" in result["warnings"][0]


def test_vision_api_timeout_returns_warning(monkeypatch: Any) -> None:
    class FakeAsyncClient:
        def __init__(self, timeout: float) -> None:
            return None

        async def __aenter__(self) -> "FakeAsyncClient":
            return self

        async def __aexit__(self, exc_type: Any, exc: Any, traceback: Any) -> None:
            return None

        async def post(self, url: str, headers: dict[str, str], json: dict[str, Any]) -> None:
            raise httpx.ReadTimeout("upstream timed out")

    monkeypatch.setenv(main.LLM_BASE_URL_ENV, "https://api.quickrouter.ai/v1/chat/completions")
    monkeypatch.setenv(main.LLM_MODEL_ENV, "gpt-5.4-nano")
    monkeypatch.setenv(main.LLM_TIMEOUT_SECONDS_ENV, "7")

    async def fake_sleep(seconds: float) -> None:
        return None

    monkeypatch.setattr(main.httpx, "AsyncClient", FakeAsyncClient)
    monkeypatch.setattr(main.asyncio, "sleep", fake_sleep)

    response = client.post(
        "/api/parse",
        data={"text": "读取图片"},
        files={"files": ("form.png", b"image-bytes", "image/png")},
    )

    assert response.status_code == 200
    result = response.json()["data"]["result"]
    assert result["sections"][0]["fields"][0]["status"] == "uncertain"
    assert "timed out" in result["warnings"][0]
    assert "timeout=7.0s" in result["warnings"][0]
    assert "retries=2" in result["warnings"][0]


def test_parse_persists_result_when_database_is_configured(monkeypatch: Any) -> None:
    captured: dict[str, Any] = {}

    def fake_save_extraction_result(request_id: str, input_text: str, result: dict[str, Any], saved_files: list[str]) -> str:
        captured["request_id"] = request_id
        captured["input_text"] = input_text
        captured["result"] = result
        captured["saved_files"] = saved_files
        return "record-123"

    monkeypatch.setattr(main, "save_extraction_result", fake_save_extraction_result)

    response = client.post(
        "/api/parse",
        data={"text": "需要保存的识别结果"},
        files={"files": ("contract.txt", b"hello", "text/plain")},
    )

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["record_id"] == "record-123"
    assert captured["input_text"] == "需要保存的识别结果"
    assert captured["result"] == body["result"]
    assert captured["saved_files"] == body["saved_files"]


def test_admin_results_returns_saved_records(monkeypatch: Any) -> None:
    records = [
        {
            "id": "record-123",
            "request_id": "request-123",
            "input_text": "发票",
            "result_json": {"document_info": {"title": "发票"}, "sections": [], "raw_text": "", "warnings": []},
            "saved_files": [],
            "created_at": "2026-06-24 00:00:00+00",
        }
    ]

    monkeypatch.setattr(main, "get_database_url", lambda: "postgresql://example")
    monkeypatch.setattr(main, "list_extraction_results", lambda limit=100: records)

    response = client.get("/api/admin/results")

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["database_enabled"] is True
    assert body["items"] == records


def test_admin_result_detail_returns_record(monkeypatch: Any) -> None:
    record = {
        "id": "record-123",
        "request_id": "request-123",
        "input_text": "表格",
        "result_json": {"document_info": {"title": "表格"}, "sections": [], "raw_text": "", "warnings": []},
        "saved_files": [],
        "created_at": "2026-06-24 00:00:00+00",
    }

    monkeypatch.setattr(main, "get_extraction_result", lambda record_id: record if record_id == "record-123" else None)

    response = client.get("/api/admin/results/record-123")

    assert response.status_code == 200
    assert response.json()["data"] == record


def test_static_files_do_not_return_not_modified() -> None:
    static_files = main.NoCacheStaticFiles(directory=".")

    assert static_files.is_not_modified(
        Headers({"etag": "example"}),
        Headers({"if-none-match": "example"}),
    ) is False


def test_favicon_ico_returns_icon_response() -> None:
    response = client.get("/favicon.ico")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("image/svg+xml")
    assert response.headers["cache-control"] == "no-store"
    assert b"<svg" in response.content


def test_admin_results_can_search_database_fields(monkeypatch: Any) -> None:
    records = [
        {
            "id": "record-123",
            "request_id": "request-123",
            "input_text": "巡检",
            "result_json": {"fields": [{"key": "temperature", "value": "23 °C"}]},
            "saved_files": [],
            "created_at": "2026-06-25 00:00:00+00",
        }
    ]
    captured: dict[str, Any] = {}

    def fake_search(query: str, limit: int = 100) -> list[dict[str, Any]]:
        captured["query"] = query
        captured["limit"] = limit
        return records

    monkeypatch.setattr(main, "get_database_url", lambda: "postgresql://example")
    monkeypatch.setattr(main, "search_extraction_results", fake_search)

    response = client.get("/api/admin/results?query=temperature&limit=25")

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["database_enabled"] is True
    assert body["query"] == "temperature"
    assert body["items"] == records
    assert captured == {"query": "temperature", "limit": 25}


def test_admin_result_update_delete_endpoints(monkeypatch: Any) -> None:
    records: dict[str, dict[str, Any]] = {
        "record-123": {
            "id": "record-123",
            "request_id": "request-123",
            "input_text": "原始",
            "result_json": {"fields": [{"key": "title", "value": "原始"}]},
            "saved_files": [],
            "created_at": "2026-06-25 00:00:00+00",
        }
    }

    def fake_update(record_id: str, input_text: str, result: dict[str, Any], saved_files: list[str] | None = None) -> bool:
        if record_id not in records:
            return False
        records[record_id].update(
            {
                "input_text": input_text,
                "result_json": result,
                "saved_files": saved_files or [],
            }
        )
        return True

    def fake_delete(record_id: str) -> bool:
        return records.pop(record_id, None) is not None

    monkeypatch.setattr(main, "update_extraction_result", fake_update)
    monkeypatch.setattr(main, "delete_extraction_result", fake_delete)
    monkeypatch.setattr(main, "get_extraction_result", lambda record_id: records.get(record_id))

    update_response = client.put(
        "/api/admin/results/record-123",
        json={"input_text": "更新", "result_json": {"fields": [{"key": "title", "value": "更新"}]}},
    )

    assert update_response.status_code == 200
    assert update_response.json()["data"]["input_text"] == "更新"
    assert records["record-123"]["result_json"]["fields"][0]["value"] == "更新"

    delete_response = client.delete("/api/admin/results/record-123")

    assert delete_response.status_code == 200
    assert delete_response.json()["data"] == {"id": "record-123"}
    assert records == {}


def test_admin_result_file_downloads_saved_upload(monkeypatch: Any) -> None:
    request_dir = main.UPLOAD_DIR / "test-request"
    request_dir.mkdir(parents=True, exist_ok=True)
    image_path = request_dir / "form.png"
    image_path.write_bytes(b"image-bytes")

    record = {
        "id": "record-123",
        "request_id": "request-123",
        "input_text": "表格",
        "result_json": {"document_info": {"title": "表格"}, "sections": [], "raw_text": "", "warnings": []},
        "saved_files": [str(image_path)],
        "created_at": "2026-06-24 00:00:00+00",
    }

    monkeypatch.setattr(main, "get_extraction_result", lambda record_id: record if record_id == "record-123" else None)

    response = client.get("/api/admin/results/record-123/files/0")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("image/png")
    assert response.content == b"image-bytes"


def test_export_excel_uses_field_table_columns() -> None:
    payload = {
        "data": {
            "document_info": {"title": "巡检表"},
            "sections": [
                {
                    "section_name": "基础信息",
                    "fields": [
                        {
                            "field_name": "温度",
                            "field_value": "23 °C",
                            "status": "filled",
                            "source_hint": "第一行",
                        }
                    ],
                }
            ],
            "raw_text": "不需要展示",
            "warnings": [],
        }
    }

    response = client.post("/api/export/excel", json=payload)

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0] == ("section", "field", "value", "status", "source")
    assert rows[1] == ("基础信息", "温度", "23 °C", "filled", "第一行")


def test_export_json_downloads_file() -> None:
    payload = {"data": {"summary": "demo", "fields": [{"key": "title", "value": "demo"}]}}
    response = client.post("/api/export/json", json=payload)

    assert response.status_code == 200
    assert response.headers["content-disposition"].startswith('attachment; filename="result.json"')
    assert json.loads(response.content) == payload["data"]
